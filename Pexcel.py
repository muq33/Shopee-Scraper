from openpyxl import Workbook
from openpyxl import load_workbook
import os.path

def make_book(num_products):
    columns = ['Nome do produto', 'Preço', 'Desconto', 'Link']
    wb = Workbook()
    
    file_name = 'Shopee.xlsx'
    wb_sheet = wb['Sheet']
    wb_sheet.title = 'Produto 1'
    if num_products == 1:
        wb.create_sheet('Variação Produto 1')
    else:
        for i in range(2, num_products+1):
            wb.create_sheet(f'Produto {i}')
            wb.create_sheet(f'Variação Produto {i}')
    
    for i in range(1,num_products+1):
        for j in range(1,5):
            wb[f'Produto {i}'].cell(column = j, row = 1, value = columns[j-1])
    wb.save(filename = file_name)
    return None

def write_data(list_products):
    if os.path.exists('Shopee.xlsx'):
        wb = load_workbook('Shopee.xlsx')
        for i in range(len(list_products)):
            last_value = wb[f'Variação Produto {i+1}'].max_column
            for j in range(len(list_products[i])):
                for k in range(4):
                    wb[f'Produto {i+1}'].cell(column = k+1, row = j+2, value = list_products[i][j][k])
                               
                wb[f'Variação Produto {i+1}'].cell(column = last_value+1, row = j+2, value = list_products[i][j][1])
        wb.save(filename = 'Shopee.xlsx')
    else:
        make_book(len(list_products))
        wb = load_workbook('Shopee.xlsx')
        for i in range(len(list_products)):
            last_value = wb[f'Variação Produto {i+1}'].max_column
            for j in range(len(list_products[i])):
                for k in range(4):
                     wb[f'Produto {i+1}'].cell(column = k+1, row = j+2, value = list_products[i][j][k])
                  
                wb[f'Variação Produto {i+1}'].cell(column = last_value+1, row = j+2, value = list_products[i][j][1])
        wb.save(filename = 'Shopee.xlsx')
    return None
    #except:
        #1 #make_book(n_prod)